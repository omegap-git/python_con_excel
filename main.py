from openpyxl import Workbook

def crear_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ejemplo"

    ws["A1"] = "Nombre"
    ws["B1"] = "Edad"

    ws.append(["Juan", 30])
    ws.append(["Ana", 25])

    wb.save("ejemplo.xlsx")

if __name__ == "__main__":
    crear_excel()